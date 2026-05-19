from pathlib import Path
from openpyxl import load_workbook
import re
import hashlib

repo = Path('/Users/user/Frontend-Development-')
wb_path = repo / 'Frontend_Development_Practice_Bank.xlsx'
wb = load_workbook(wb_path)

owner = 'Nithya-sri-14'
repo_name = 'Frontend-Development-'
branch = 'main'
img_root = repo / 'expected-pictures'
img_root.mkdir(parents=True, exist_ok=True)

def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def pick_palette(topic):
    m = {
        'HTML': ('#0f766e','#f0fdfa'),
        'Semantic HTML': ('#0369a1','#f0f9ff'),
        'Forms': ('#1d4ed8','#eff6ff'),
        'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),
        'Box Model': ('#b45309','#fff7ed'),
        'Flexbox': ('#0e7490','#ecfeff'),
        'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),
        'Responsive Design': ('#9a3412','#fff7ed'),
    }
    return m.get(topic, ('#1f2937','#f8fafc'))

def section_labels(topic, title):
    t = title.lower()
    if 'portfolio' in t:
        return ['Nav', 'Hero', 'About', 'Projects', 'Contact']
    if topic == 'Forms':
        return ['Form Title','Personal Fields','Selection Group','Validation Notes','Submit Area']
    if topic == 'Tables':
        return ['Table Caption','Header Row','Data Region','Summary Row','Notes']
    if topic == 'Semantic HTML':
        return ['Header','Main Article','Aside','Figure','Footer']
    if topic == 'Flexbox':
        return ['Toolbar','Primary Row','Actions','Wrapped Items','Footer Row']
    if topic == 'Grid':
        return ['Top Strip','Sidebar','Grid A','Grid B','Bottom Strip']
    if topic == 'Media Queries':
        return ['Desktop View','Tablet View','Mobile View','Breakpoint Rules','Adaptive Block']
    if topic == 'Responsive Design':
        return ['Header','Fluid Hero','Adaptive Cards','Content Flow','Footer']
    if topic == 'Box Model':
        return ['Margin','Border','Padding','Content','Spacing']
    if topic == 'CSS':
        return ['Theme','Typography','Buttons','Cards','States']
    return ['Header','Section A','Section B','Section C','Footer']

def unique_layout(seed):
    # deterministic unique layout from hash; returns list of 5 rectangles
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    pattern = h % 6
    if pattern == 0:
        return [(80,190,1120,90),(80,295,550,150),(650,295,550,150),(80,460,760,150),(860,460,340,150)]
    if pattern == 1:
        return [(80,190,1120,90),(80,295,360,315),(460,295,740,95),(460,405,360,205),(840,405,360,205)]
    if pattern == 2:
        return [(80,190,1120,90),(80,295,1120,120),(80,430,360,180),(460,430,360,180),(840,430,360,180)]
    if pattern == 3:
        return [(80,190,360,200),(460,190,360,200),(840,190,360,200),(80,410,540,200),(640,410,560,200)]
    if pattern == 4:
        return [(80,190,760,120),(860,190,340,120),(80,330,360,280),(460,330,360,280),(840,330,360,280)]
    return [(80,190,540,200),(640,190,560,200),(80,410,1120,90),(80,520,550,90),(650,520,550,90)]

def make_svg(path: Path, topic: str, title: str):
    accent, bg = pick_palette(topic)
    labels = section_labels(topic, title)
    rects = unique_layout(topic + '|' + title)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    body = []
    for i, (x,y,w,h) in enumerate(rects):
        body.append(f"<rect x='{x}' y='{y}' width='{w}' height='{h}' rx='12' fill='{fills[i]}'/>")
        body.append(f"<text x='{x+14}' y='{y+34}' font-family='Arial, sans-serif' font-size='24' font-weight='700' fill='#111827'>{labels[i]}</text>")

    svg = f"""<svg xmlns='http://www.w3.org/2000/svg' width='1280' height='720' viewBox='0 0 1280 720'>
  <rect width='1280' height='720' fill='{bg}'/>
  <rect x='36' y='36' width='1208' height='648' rx='20' fill='white' stroke='{accent}' stroke-width='5'/>
  <rect x='60' y='60' width='1160' height='92' rx='12' fill='white'/>
  <text x='78' y='98' font-family='Arial, sans-serif' font-size='30' font-weight='800' fill='{accent}'>{title}</text>
  <text x='78' y='132' font-family='Arial, sans-serif' font-size='20' fill='#334155'>Expected UI Reference</text>
  {''.join(body)}
  <text x='80' y='660' font-family='Arial, sans-serif' font-size='20' fill='#0f172a'>Recreate this exact final interface.</text>
</svg>"""
    path.write_text(svg, encoding='utf-8')

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        topic = ws.cell(r, 3).value
        title = ws.cell(r, 4).value

        d = img_root / slugify(sheet) / slugify(topic)
        d.mkdir(parents=True, exist_ok=True)
        fname = f"q{r-1:03d}-{slugify(title)[:90]}.svg"
        p = d / fname
        make_svg(p, topic, title)

        rel = p.relative_to(repo).as_posix()
        ws.cell(r, 9).value = f"https://raw.githubusercontent.com/{owner}/{repo_name}/{branch}/{rel}"

wb.save(wb_path)
print('Updated unique, topic-relevant picture outputs without level labels.')
