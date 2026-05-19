from pathlib import Path
from openpyxl import load_workbook
import re

repo = Path('/Users/user/Frontend-Development-')
wb_path = repo / 'Frontend_Development_Practice_Bank.xlsx'
wb = load_workbook(wb_path)

owner = 'Nithya-sri-14'
repo_name = 'Frontend-Development-'
# Renderable public URL for HTML in GitHub repo
base_preview = f'https://raw.githack.com/{owner}/{repo_name}/main'

out_root = repo / 'expected-html'
out_root.mkdir(parents=True, exist_ok=True)

def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def make_constraints(level, topic):
    common = {
        'Beginner': 'Match the expected output exactly for layout, spacing, typography, and colors. Use only HTML and CSS (no JavaScript).',
        'Intermediate': 'Match the expected output exactly across mobile and desktop views. Use semantic HTML, reusable CSS classes, and visible focus states.',
        'Hard': 'Match the expected output exactly across mobile, tablet, and desktop with no layout breakage. Ensure accessibility-compliant contrast, focus visibility, and maintainable CSS architecture.'
    }[level]

    topic_rules = {
        'HTML': 'Structure must use correct document hierarchy and meaningful sectioning elements.',
        'Semantic HTML': 'Use appropriate landmark and semantic elements; avoid generic div wrappers where semantic tags apply.',
        'Forms': 'All controls must have proper labels, grouping, and validation-friendly structure.',
        'Tables': 'Use proper table semantics including caption, thead/tbody, and scoped headers where needed.',
        'CSS': 'Use class-based styling with clean selector strategy and no inline styles.',
        'Box Model': 'Spacing and sizing must be implemented using margin, padding, border, and box-sizing intentionally.',
        'Flexbox': 'Primary alignment and distribution must be solved with Flexbox properties.',
        'Grid': 'Main layout must be solved with CSS Grid tracks/areas, not float hacks.',
        'Media Queries': 'Use breakpoint rules to reproduce the specified layout changes exactly.',
        'Responsive Design': 'Use fluid/adaptive techniques so UI remains consistent across viewport sizes.'
    }[topic]

    return f"{common} {topic_rules}"

def make_html(level, topic, title, qid):
    palette = {
        'Beginner': ('#0f766e', '#f0fdfa'),
        'Intermediate': ('#1d4ed8', '#eff6ff'),
        'Hard': ('#b45309', '#fff7ed')
    }
    accent, pagebg = palette[level]
    v = qid % 5
    blocks = {
        0: """
      <section class='row a'></section>
      <section class='two'>
        <article class='card b'></article>
        <article class='card c'></article>
      </section>
""",
        1: """
      <section class='two'>
        <article class='sidebar'></article>
        <div class='stack'>
          <article class='card a'></article>
          <section class='two'>
            <article class='card b'></article>
            <article class='card c'></article>
          </section>
        </div>
      </section>
""",
        2: """
      <section class='four'>
        <article class='card a'></article>
        <article class='card b'></article>
        <article class='card c'></article>
        <article class='card d'></article>
      </section>
      <section class='row e'></section>
""",
        3: """
      <section class='row f'></section>
      <section class='three'>
        <article class='card a'></article>
        <article class='card b'></article>
        <article class='card c'></article>
      </section>
""",
        4: """
      <section class='two'>
        <article class='hero'></article>
        <div class='stack'>
          <article class='card a'></article>
          <article class='card b'></article>
          <article class='card c'></article>
        </div>
      </section>
""",
    }[v]

    html = f"""<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8'>
  <meta name='viewport' content='width=device-width, initial-scale=1'>
  <title>{title} - Expected Output</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; background: {pagebg}; color: #0f172a; }}
    .wrap {{ max-width: 1160px; margin: 28px auto; padding: 24px; background: #fff; border: 4px solid {accent}; border-radius: 16px; }}
    h1 {{ margin: 0 0 10px; font-size: 32px; color: {accent}; }}
    p.meta {{ margin: 0 0 18px; font-size: 18px; }}
    .grid {{ display: grid; gap: 16px; }}
    .row {{ height: 92px; border-radius: 12px; }}
    .two {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
    .three {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 16px; }}
    .four {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 16px; }}
    .stack {{ display: grid; gap: 16px; }}
    .card, .hero, .sidebar {{ height: 160px; border-radius: 12px; }}
    .hero {{ height: 336px; }}
    .sidebar {{ height: 336px; }}
    .a {{ background: #dbeafe; }}
    .b {{ background: #e2e8f0; }}
    .c {{ background: #fef3c7; }}
    .d {{ background: #ede9fe; }}
    .e {{ background: #f1f5f9; }}
    .f {{ background: #ecfeff; }}
    .cta {{ margin-top: 18px; display: inline-block; padding: 12px 18px; border-radius: 10px; background: {accent}; color: #fff; font-weight: 700; }}
    @media (max-width: 900px) {{
      .two, .three, .four {{ grid-template-columns: 1fr; }}
      .hero, .sidebar {{ height: 180px; }}
      h1 {{ font-size: 26px; }}
    }}
  </style>
</head>
<body>
  <main class='wrap'>
    <h1>{title}</h1>
    <p class='meta'>{level} • {topic} • Exact expected output reference</p>
    <section class='grid'>
      {blocks}
    </section>
    <span class='cta'>Recreate this exact UI</span>
  </main>
</body>
</html>
"""
    return html

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        level = ws.cell(r, 2).value
        topic = ws.cell(r, 3).value
        title = ws.cell(r, 4).value
        qid = r - 1

        file_slug = slugify(title)[:90]
        html_dir = out_root / slugify(level) / slugify(topic)
        html_dir.mkdir(parents=True, exist_ok=True)
        html_path = html_dir / f'q{qid:03d}-{file_slug}.html'
        html_path.write_text(make_html(level, topic, title, qid), encoding='utf-8')

        rel = html_path.relative_to(repo).as_posix()
        ws.cell(r, 9).value = f'{base_preview}/{rel}'
        ws.cell(r, 8).value = make_constraints(level, topic)

wb.save(wb_path)
print('Updated workbook with exact runnable HTML output links and revised constraints.')
