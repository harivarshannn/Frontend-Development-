from pathlib import Path
from openpyxl import Workbook
import hashlib
import re

repo = Path('/Users/user/Frontend-Development-')
out_xlsx = repo / 'Frontend_Development_Practice_Bank.xlsx'
solutions_root = repo / 'solutions'
solutions_root.mkdir(parents=True, exist_ok=True)

owner = 'Nithya-sri-14'
repo_name = 'Frontend-Development-'
preview_base = f'https://raw.githack.com/{owner}/{repo_name}/main'

levels = ['Beginner', 'Intermediate', 'Hard']
topics = [
    'HTML', 'Semantic HTML', 'Forms', 'Tables', 'CSS',
    'Box Model', 'Flexbox', 'Grid', 'Media Queries', 'Responsive Design'
]

language = 'Frontend Development'

# unique theme words per level so no overlap
level_themes = {
    'Beginner': ['Starter', 'Foundation', 'Core', 'Intro', 'Basic', 'Primary', 'Early', 'First', 'Simple', 'Essential'],
    'Intermediate': ['Applied', 'Workflow', 'Module', 'System', 'Pattern', 'Scenario', 'Practice', 'Iteration', 'Refinement', 'Studio'],
    'Hard': ['Production', 'Enterprise', 'Scale', 'Robust', 'Advanced', 'Architecture', 'Audit', 'Optimization', 'Resilience', 'Mastery'],
}

# 30 distinct intents per topic (shared structure, but final titles include level-specific theme + unique id)
topic_intents = {
    'HTML': ['Portfolio Landing', 'Club Portal', 'Startup Story Page', 'Travel Journal Intro', 'Product Launch Page', 'Learning Hub Homepage', 'Event Brief Page', 'Nonprofit Mission Page', 'Fitness Coach Profile', 'Book Review Portal', 'Research Announcement Page', 'Workshop Information Page', 'Town Bulletin Page', 'Internship Intro Site', 'Career Expo Page', 'Community Update Page', 'Festival Overview Page', 'Restaurant Welcome Page', 'Course Catalog Intro', 'Podcast Episode Page', 'School Notice Page', 'Public Service Update', 'Gallery Showcase Page', 'Author Bio Page', 'Company Values Page', 'Support Center Intro', 'Healthcare Info Page', 'Music Showcase Page', 'Tourism Intro Page', 'Developer Conference Page'],
    'Semantic HTML': ['Editorial News Story', 'Policy Explanation Article', 'Research Digest', 'Accessibility-first Guide', 'Civic Update Article', 'Medical Advisory Article', 'Environmental Report', 'Cultural Feature Story', 'Open Source Changelog', 'Case Study Narrative', 'Interview Transcript Page', 'Magazine Longread', 'Standards Overview Article', 'Learning Roadmap Article', 'Charity Impact Story', 'Travel Insight Article', 'Sports Analysis Piece', 'Film Critique Article', 'Technology Brief', 'Public Program Summary', 'Museum Exhibit Story', 'City Bulletin Narrative', 'Education Reform Story', 'Innovation Report', 'Agriculture Brief', 'Security Incident Summary', 'Library Program Report', 'Conference Recap', 'Community Program Story', 'Startup Milestone Story'],
    'Forms': ['Admission Application Form', 'Support Ticket Form', 'Appointment Booking Form', 'Scholarship Registration Form', 'Event Speaker Form', 'Internship Submission Form', 'Volunteer Enrollment Form', 'Travel Reimbursement Form', 'Lab Test Request Form', 'Feedback Survey Form', 'Partnership Request Form', 'KYC Update Form', 'Insurance Claim Form', 'Subscription Upgrade Form', 'Contest Entry Form', 'Course Enrollment Form', 'Restaurant Booking Form', 'Rental Inquiry Form', 'Project Brief Form', 'Conference Registration Form', 'Mentor Matching Form', 'Complaint Intake Form', 'Parent Meeting Form', 'Catering Inquiry Form', 'Trial Activation Form', 'Research Consent Form', 'Leave Request Form', 'Job Referral Form', 'Profile Update Form', 'Payment Details Form'],
    'Tables': ['Grade Register Table', 'Inventory Ledger Table', 'Sales Performance Table', 'Shift Roster Table', 'Flight Timetable Table', 'Budget Summary Table', 'Order Audit Table', 'Attendance Log Table', 'Campaign Metrics Table', 'Sprint Tracking Table', 'Revenue Comparison Table', 'Course Schedule Table', 'Subscription Matrix Table', 'Traffic Summary Table', 'Supplier Tracker Table', 'Experiment Log Table', 'Incident Timeline Table', 'Helpdesk Resolution Table', 'Payroll Summary Table', 'League Standings Table', 'Seating Allocation Table', 'Usage Report Table', 'Conversion Dashboard Table', 'Defect Report Table', 'Yield Comparison Table', 'Transit Route Table', 'Training Matrix Table', 'NPS Breakdown Table', 'Library Circulation Table', 'Milestone Tracker Table'],
    'CSS': ['Theme Styling Challenge', 'Typography System Challenge', 'Button State Challenge', 'Card Skin Challenge', 'Navigation Style Challenge', 'Promo Banner Challenge', 'Alert Style Challenge', 'Tag Color Challenge', 'Sidebar Theme Challenge', 'Footer Hierarchy Challenge', 'Timeline Visual Challenge', 'Widget Appearance Challenge', 'Pricing Card Challenge', 'Hero Section Challenge', 'Form State Challenge', 'Panel Depth Challenge', 'CTA Treatment Challenge', 'Tab Styling Challenge', 'Badge Priority Challenge', 'Review Quote Challenge', 'Section Divider Challenge', 'Search Panel Challenge', 'Color Token Challenge', 'Microcopy Tone Challenge', 'Hover Atmosphere Challenge', 'Loading State Challenge', 'Notification Style Challenge', 'Feature Emphasis Challenge', 'Utility Theme Challenge', 'Article Rhythm Challenge'],
    'Box Model': ['Margin Strategy Task', 'Padding Strategy Task', 'Border Framing Task', 'Content Width Task', 'Spacing Rhythm Task', 'Card Boundary Task', 'Toolbar Gap Task', 'Section Shell Task', 'Panel Sizing Task', 'List Density Task', 'Caption Wrapper Task', 'Avatar Meta Task', 'CTA Breathing Task', 'Table Cell Spacing Task', 'Comment Rhythm Task', 'Modal Framing Task', 'Sidebar Item Sizing Task', 'Footer Block Sizing Task', 'Article Width Guard Task', 'Pricing Edge Task', 'Notification Container Task', 'Timeline Node Task', 'Action Bar Task', 'Profile Frame Task', 'Checkout Boundary Task', 'Stack Offset Task', 'Grid Tile Box Task', 'Header Body Gap Task', 'Form Field Spacing Task', 'FAQ Panel Spacing Task'],
    'Flexbox': ['Navigation Distribution Task', 'Toolbar Alignment Task', 'Card Row Wrapping Task', 'Action Cluster Task', 'Header Justification Task', 'Profile Row Alignment Task', 'Icon Label Pairing Task', 'KPI Row Balance Task', 'Tag Wrap Behavior Task', 'Footer Link Alignment Task', 'Media Object Task', 'Result Row Task', 'Button Cluster Task', 'Menu Toggle Task', 'Notification Row Task', 'Filter Bar Task', 'Timeline Row Task', 'Chat Row Task', 'Pricing Row Task', 'Gallery Caption Task', 'Task Item Alignment Task', 'Hero CTA Group Task', 'Sidebar Menu Task', 'Checkout Totals Task', 'Baseline Alignment Task', 'Sticky Footer Task', 'Comment Header Task', 'Control Strip Task', 'Card Deck Equalization Task', 'Mobile Header Shift Task'],
    'Grid': ['Dashboard Grid Task', 'Catalog Grid Task', 'Portfolio Grid Task', 'Magazine Grid Task', 'Service Grid Task', 'Pricing Grid Task', 'Feature Matrix Grid Task', 'Sidebar Content Grid Task', 'Planner Grid Task', 'Knowledge Grid Task', 'Gallery Proportion Grid Task', 'Campaign Grid Task', 'Widget Board Grid Task', 'Operations Grid Task', 'Calendar Grid Task', 'Team Directory Grid Task', 'Case Study Grid Task', 'Landing Feature Grid Task', 'Admin Console Grid Task', 'Department Overview Grid Task', 'Offer Comparison Grid Task', 'Learning Module Grid Task', 'Research Summary Grid Task', 'Marketplace Grid Task', 'Stats Table Grid Task', 'Product Shelf Grid Task', 'News Digest Grid Task', 'KPI Snapshot Grid Task', 'Onboarding Grid Task', 'Schedule Grid Task'],
    'Media Queries': ['Breakpoint Navigation Shift', 'Card Reflow Shift', 'Typography Scale Shift', 'Form Layout Shift', 'Sidebar Collapse Shift', 'Hero Ratio Shift', 'Button Wrap Shift', 'Table-to-Card Shift', 'Gallery Column Shift', 'Footer Stack Shift', 'Sticky Header Shift', 'Section Spacing Shift', 'Banner Height Shift', 'CTA Position Shift', 'Grid Area Shift', 'Image Treatment Shift', 'Input Width Shift', 'Pricing Order Shift', 'Visibility Shift', 'Dashboard Block Shift', 'Two-pane Shift', 'Density Shift', 'Metric Rearrangement Shift', 'Timeline Simplification Shift', 'Control Panel Shift', 'Search Layout Shift', 'Feature Rebalance Shift', 'Copy Scaling Shift', 'Hit Area Shift', 'Layering Shift'],
    'Responsive Design': ['Multi-device Landing Task', 'Adaptive Education Task', 'Fluid Clinic Task', 'Responsive Restaurant Task', 'Travel Adaptive Task', 'Portfolio Fluid Task', 'Community Adaptive Task', 'Event Responsive Task', 'Startup Adaptive Task', 'News Multi-breakpoint Task', 'Dashboard Adaptive Task', 'Booking Responsive Task', 'Magazine Responsive Task', 'Campaign Responsive Task', 'Ecommerce Adaptive Task', 'Help Center Responsive Task', 'Docs Adaptive Task', 'Course Card Responsive Task', 'Festival Adaptive Task', 'Research Portal Responsive Task', 'Banking Adaptive Task', 'Recruitment Responsive Task', 'Public Service Responsive Task', 'Fitness Adaptive Task', 'Agriculture Responsive Task', 'Movie Review Responsive Task', 'Real Estate Adaptive Task', 'Logistics Responsive Task', 'Tourism Adaptive Task', 'Hackathon Responsive Task'],
}

constraints_topic = {
    'HTML': 'Use correct document hierarchy and semantic sectioning for all page regions.',
    'Semantic HTML': 'Use landmark and semantic tags appropriately; avoid non-semantic wrappers where semantic tags apply.',
    'Forms': 'All form controls must have explicit labels and logical grouping.',
    'Tables': 'Use proper table structure (`caption`, `thead`, `tbody`, and scoped headers).',
    'CSS': 'Use class-based styling with clear, reusable selectors and no inline styles.',
    'Box Model': 'Implement spacing using margin, padding, border, and box-sizing intentionally.',
    'Flexbox': 'Primary alignment/distribution must be solved using Flexbox properties.',
    'Grid': 'Primary layout must be solved using CSS Grid tracks/areas.',
    'Media Queries': 'Use breakpoints to reproduce the required layout transitions exactly.',
    'Responsive Design': 'Ensure fluid/adaptive behavior across viewport sizes without layout breakage.',
}

constraints_level = {
    'Beginner': 'Use only HTML and CSS. Match spacing, typography, and alignment exactly.',
    'Intermediate': 'Use mobile-first CSS and at least 2 breakpoints with reusable classes and visible focus states.',
    'Hard': 'Support mobile/tablet/desktop with production-grade maintainable CSS and consistent accessibility.',
}


def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def palette(topic):
    return {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }[topic]

def unique_layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 6
    if p == 0:
        return [(80,190,1120,90),(80,295,550,150),(650,295,550,150),(80,460,760,150),(860,460,340,150)]
    if p == 1:
        return [(80,190,1120,90),(80,295,360,315),(460,295,740,95),(460,405,360,205),(840,405,360,205)]
    if p == 2:
        return [(80,190,1120,90),(80,295,1120,120),(80,430,360,180),(460,430,360,180),(840,430,360,180)]
    if p == 3:
        return [(80,190,360,200),(460,190,360,200),(840,190,360,200),(80,410,540,200),(640,410,560,200)]
    if p == 4:
        return [(80,190,760,120),(860,190,340,120),(80,330,360,280),(460,330,360,280),(840,330,360,280)]
    return [(80,190,540,200),(640,190,560,200),(80,410,1120,90),(80,520,550,90),(650,520,550,90)]

def labels_for(topic, title):
    if topic == 'Forms':
        return ['Form Header','Input Group','Choice Group','Validation Block','Submit Panel']
    if topic == 'Tables':
        return ['Caption Area','Header Row','Data Region','Summary Row','Legend']
    if topic == 'Semantic HTML':
        return ['Header','Main','Aside','Figure','Footer']
    if topic == 'Flexbox':
        return ['Top Bar','Primary Flex Row','Action Group','Wrapped Items','Bottom Bar']
    if topic == 'Grid':
        return ['Top Strip','Left Grid','Right Grid','Lower Grid','Footer Strip']
    if topic == 'Media Queries':
        return ['Desktop State','Tablet State','Mobile State','Breakpoint Notes','Adaptive Block']
    if topic == 'Responsive Design':
        return ['Responsive Header','Fluid Hero','Adaptive Cards','Flexible Content','Responsive Footer']
    if topic == 'Box Model':
        return ['Margin Zone','Border Zone','Padding Zone','Content Zone','Spacing System']
    if topic == 'CSS':
        return ['Theme Block','Type Block','State Block','Component Block','Token Block']
    if 'Portfolio' in title:
        return ['Nav','Hero','About','Projects','Contact']
    return ['Header','Section A','Section B','Section C','Footer']

def make_html(topic, title, seed):
    accent, bg = palette(topic)
    rects = unique_layout(seed)
    labels = labels_for(topic, title)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    blocks = []
    for i, (x,y,w,h) in enumerate(rects):
        blocks.append(f"<section class='blk' style='left:{x}px;top:{y}px;width:{w}px;height:{h}px;background:{fills[i]};'><h2>{labels[i]}</h2></section>")

    return f"""<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1' />
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; background: {bg}; color: #111827; }}
    .canvas {{ width: 1280px; min-height: 720px; margin: 0 auto; position: relative; }}
    .frame {{ position: absolute; left: 36px; top: 36px; width: 1208px; height: 648px; background: #fff; border: 5px solid {accent}; border-radius: 20px; }}
    .title {{ position: absolute; left: 78px; top: 74px; margin: 0; font-size: 30px; font-weight: 800; color: {accent}; }}
    .sub {{ position: absolute; left: 78px; top: 114px; margin: 0; font-size: 20px; color: #334155; }}
    .blk {{ position: absolute; border-radius: 12px; padding: 10px 14px; overflow: hidden; }}
    .blk h2 {{ margin: 0; font-size: 24px; font-weight: 700; color: #111827; }}
    .foot {{ position: absolute; left: 80px; top: 660px; margin: 0; font-size: 20px; color: #0f172a; }}
    @media (max-width: 1280px) {{
      .canvas {{ width: 100vw; min-height: 56.25vw; }}
      .frame, .title, .sub, .foot, .blk {{ transform-origin: top left; transform: scale(calc(100vw / 1280)); }}
    }}
  </style>
</head>
<body>
  <main class='canvas'>
    <div class='frame'></div>
    <h1 class='title'>{title}</h1>
    <p class='sub'>Expected UI Reference</p>
    {''.join(blocks)}
    <p class='foot'>Recreate this exact final interface.</p>
  </main>
</body>
</html>
"""

wb = Workbook()
wb.remove(wb.active)
headers = ['Language','Level','Topic','Title','Description','Explanation scenario','Instruction hint','Constraints','Example output']

for level in levels:
    ws = wb.create_sheet(level)
    ws.append(headers)
    q = 0
    for topic in topics:
        intents = topic_intents[topic]
        for i, intent in enumerate(intents, start=1):
            q += 1
            # globally unique title (level+theme+topic+qid)
            theme = level_themes[level][(i-1) % len(level_themes[level])]
            title = f"{theme} {intent} - {topic} Practice {level[0]}{q:03d}"

            desc = f"Build this UI from scratch using HTML and CSS for '{intent}'. The final page must match the expected output exactly."
            scen = f"This problem simulates a frontend assessment task where students must implement a {topic.lower()} solution that is visually accurate and structurally correct."
            hint = f"Start from semantic structure, map each visible block, then style layout, spacing, and typography to match the reference output precisely."
            cons = f"{constraints_level[level]} {constraints_topic[topic]}"

            level_slug = slugify(level)
            topic_slug = slugify(topic)
            file_slug = slugify(title)[:100]
            html_dir = solutions_root / level_slug / topic_slug
            html_dir.mkdir(parents=True, exist_ok=True)
            html_file = html_dir / f"q{q:03d}-{file_slug}.html"

            seed = f"{level}|{topic}|{title}"
            html_file.write_text(make_html(topic, title, seed), encoding='utf-8')
            rel = html_file.relative_to(repo).as_posix()
            out_link = f"{preview_base}/{rel}"

            ws.append([language, level, topic, title, desc, scen, hint, cons, out_link])

    for c in ['A','B','C','D','E','F','G','H','I']:
        ws.column_dimensions[c].width = 42

wb.save(out_xlsx)

# validate uniqueness globally across sheets for title/desc/hint
from openpyxl import load_workbook
check = load_workbook(out_xlsx)
titles=[]
for s in check.sheetnames:
    ws=check[s]
    for r in range(2, ws.max_row+1):
        titles.append(ws.cell(r,4).value)
print('rows',len(titles),'unique_titles',len(set(titles))==len(titles))
print(out_xlsx)
