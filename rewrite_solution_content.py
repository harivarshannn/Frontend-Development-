from pathlib import Path
from openpyxl import load_workbook
import re, hashlib

repo = Path('/Users/user/Frontend-Development-')
wb = load_workbook(repo / 'Frontend_Development_Practice_Bank.xlsx')
solutions_root = repo / 'solutions'

def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def palette(topic):
    return {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }[topic]

def unique_layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 6
    if p == 0:
        return [(80,190,1120,90),(80,295,550,150),(650,295,550,150),(80,460,760,150),(860,460,340,150)]
    if p == 1:
        return [(80,190,1120,90),(80,295,360,315),(460,295,740,95),(460,405,360,205),(840,405,360,205)]
    if p == 2:
        return [(80,190,1120,90),(80,295,1120,120),(80,430,360,180),(460,430,360,180),(840,430,360,180)]
    if p == 3:
        return [(80,190,360,200),(460,190,360,200),(840,190,360,200),(80,410,540,200),(640,410,560,200)]
    if p == 4:
        return [(80,190,760,120),(860,190,340,120),(80,330,360,280),(460,330,360,280),(840,330,360,280)]
    return [(80,190,540,200),(640,190,560,200),(80,410,1120,90),(80,520,550,90),(650,520,550,90)]

def core_title(t):
    return re.sub(r'\s-\s[^-]*Practice\s[BIH]\d{3}$', '', t).strip()

def tokens(t):
    return [w for w in re.split(r'[^A-Za-z0-9]+', t) if w]

def section_content(topic, title):
    c = core_title(title)
    w = tokens(c)
    key = ' '.join(w[:3]) if w else c
    if topic == 'Forms':
        return [
            ('Form Heading', f'{c} form'),
            ('User Details', 'Name, email, and phone input fields with aligned labels'),
            ('Options', 'Dropdown, radio group, and checkbox agreement area'),
            ('Validation', 'Inline helper text and required-field indicators'),
            ('Submission', 'Primary submit button with secondary reset action')
        ]
    if topic == 'Tables':
        return [
            ('Dataset Title', f'{c} data table'),
            ('Columns', 'Item | Category | Status | Value'),
            ('Rows', f'Populate realistic rows for {key.lower()} metrics'),
            ('Summary', 'Totals/average row at bottom with emphasis style'),
            ('Legend', 'Color/label note explaining table status values')
        ]
    if topic == 'Semantic HTML':
        return [
            ('Header Landmark', f'{c} page header and metadata'),
            ('Main Article', f'Primary narrative content for {key.lower()}'),
            ('Aside Context', 'Supporting notes, references, or quick facts'),
            ('Figure Block', 'Visual + caption area tied to article topic'),
            ('Footer Landmark', 'Author/date/source info using semantic footer')
        ]
    if topic == 'Media Queries':
        return [
            ('Desktop State', f'Wide layout for {c} with multi-column structure'),
            ('Tablet State', 'Reflow to balanced two-column arrangement'),
            ('Mobile State', 'Stacked single-column layout with full-width blocks'),
            ('Breakpoint Rules', 'Explicit breakpoints controlling spacing and order'),
            ('Adaptive Behavior', 'Typography and controls scaled per viewport')
        ]
    if topic == 'Responsive Design':
        return [
            ('Adaptive Header', f'Navigation and branding for {c}'),
            ('Fluid Hero', 'Scales proportionally with viewport width'),
            ('Content Cards', f'Responsive cards presenting {key.lower()} info'),
            ('Flexible Body', 'Content flow adjusts without overlap or clipping'),
            ('Responsive Footer', 'Footer links wrap cleanly on smaller screens')
        ]
    if topic == 'Flexbox':
        return [
            ('Primary Flex Row', f'Horizontal arrangement for {c}'),
            ('Alignment Rules', 'Use justify-content and align-items appropriately'),
            ('Wrapped Group', 'Items wrap with consistent row/column gaps'),
            ('Action Cluster', 'Buttons/actions aligned with spacing discipline'),
            ('Final Row', 'Footer/action row aligned with flex distribution')
        ]
    if topic == 'Grid':
        return [
            ('Grid Header', f'Top region for {c}'),
            ('Grid Area A', 'Named area for primary information block'),
            ('Grid Area B', 'Secondary content area with equal rhythm'),
            ('Grid Area C', 'Supplementary panel or metrics module'),
            ('Grid Footer', 'Bottom strip aligned to grid template')
        ]
    if topic == 'Box Model':
        return [
            ('Outer Margin', f'Whitespace around {c} container'),
            ('Border Frame', 'Visible boundary showing edge thickness'),
            ('Inner Padding', 'Content breathing space inside block'),
            ('Content Region', 'Text/content area with controlled width'),
            ('Spacing Scale', 'Consistent 8/12/16/24 spacing application')
        ]
    if topic == 'CSS':
        return [
            ('Theme Layer', f'Color and style direction for {c}'),
            ('Typography Layer', 'Heading/body hierarchy and readable contrast'),
            ('Component Layer', 'Card, button, and badge styling system'),
            ('State Layer', 'Hover/focus/active state visuals'),
            ('Token Layer', 'Reusable spacing/color tokens in classes')
        ]
    return [
        ('Header', f'{c} overview section'),
        ('Primary', f'Core content block for {key.lower()}'),
        ('Secondary', 'Supporting section with related information'),
        ('Highlights', 'Key points presented with clear visual hierarchy'),
        ('Footer', 'Closing area with CTA or links')
    ]

def make_html(topic, title, seed):
    accent, bg = palette(topic)
    rects = unique_layout(seed)
    sections = section_content(topic, title)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    blocks = []
    for i, (x,y,w,h) in enumerate(rects):
        h1, p1 = sections[i]
        blocks.append(
            f"<section class='blk' style='left:{x}px;top:{y}px;width:{w}px;height:{h}px;background:{fills[i]};'>"
            f"<h2>{h1}</h2><p>{p1}</p></section>"
        )

    return f"""<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1' />
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; background: {bg}; color: #111827; }}
    .canvas {{ width: 1280px; min-height: 720px; margin: 0 auto; position: relative; }}
    .frame {{ position: absolute; left: 36px; top: 36px; width: 1208px; height: 648px; background: #fff; border: 5px solid {accent}; border-radius: 20px; }}
    .title {{ position: absolute; left: 78px; top: 74px; margin: 0; font-size: 30px; font-weight: 800; color: {accent}; }}
    .sub {{ position: absolute; left: 78px; top: 114px; margin: 0; font-size: 20px; color: #334155; }}
    .blk {{ position: absolute; border-radius: 12px; padding: 10px 14px; overflow: hidden; }}
    .blk h2 {{ margin: 0 0 6px 0; font-size: 22px; font-weight: 700; color: #111827; }}
    .blk p {{ margin: 0; font-size: 15px; line-height: 1.35; color: #1f2937; }}
    .foot {{ position: absolute; left: 80px; top: 660px; margin: 0; font-size: 20px; color: #0f172a; }}
    @media (max-width: 1280px) {{
      .canvas {{ width: 100vw; min-height: 56.25vw; }}
      .frame, .title, .sub, .foot, .blk {{ transform-origin: top left; transform: scale(calc(100vw / 1280)); }}
    }}
  </style>
</head>
<body>
  <main class='canvas'>
    <div class='frame'></div>
    <h1 class='title'>{title}</h1>
    <p class='sub'>Expected UI Reference</p>
    {''.join(blocks)}
    <p class='foot'>Recreate this exact final interface.</p>
  </main>
</body>
</html>
"""

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        topic = ws.cell(r, 3).value
        title = ws.cell(r, 4).value
        q = r - 1
        level_slug = slugify(sheet)
        topic_slug = slugify(topic)
        file_slug = slugify(title)[:100]
        path = solutions_root / level_slug / topic_slug / f"q{q:03d}-{file_slug}.html"
        if path.exists():
            seed = f"{sheet}|{topic}|{title}"
            path.write_text(make_html(topic, title, seed), encoding='utf-8')

print('Updated all solution pages with non-random, question-specific section content.')
