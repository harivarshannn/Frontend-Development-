from pathlib import Path
from urllib.parse import urlparse
from openpyxl import load_workbook
import re, hashlib

repo = Path('/Users/user/Frontend-Development-')
wb_path = repo / 'Frontend_Development_Practice_Bank.xlsx'
wb = load_workbook(wb_path)

desc_templates = [
    "Create a complete {topic} solution for \"{core}\" and match the reference output exactly.",
    "Implement the \"{core}\" interface from scratch using {tech}, ensuring visual parity with the linked output.",
    "Your task is to code \"{core}\" as a production-style frontend screen that mirrors the expected result.",
    "Design and code \"{core}\" so the final UI structure, spacing, and hierarchy align with the reference.",
    "Translate the \"{core}\" specification into a working {topic} implementation with exact output fidelity.",
    "Develop \"{core}\" end-to-end and ensure the rendered page reflects the expected visual arrangement.",
    "Assemble the \"{core}\" page using clean HTML/CSS and reproduce the target UI without deviations.",
    "Construct \"{core}\" by following the layout cues and deliver the same final interface shown in output.",
    "Code \"{core}\" with accurate sections, alignment, and typography so it behaves like the sample result.",
    "Implement \"{core}\" as a student practice challenge and reproduce the expected output precisely."
]

scenario_templates = [
    "In a timed frontend round, candidates are evaluated on how accurately they implement \"{core}\" under layout constraints.",
    "A product team needs a faithful UI recreation of \"{core}\" before handoff; your submission is judged by visual correctness.",
    "During an LMS lab assessment, \"{core}\" is used to test structure, styling discipline, and semantic accuracy.",
    "For placement preparation, \"{core}\" acts as a design-to-code exercise where pixel-level alignment affects scoring.",
    "A mock interview challenge asks you to implement \"{core}\" while maintaining readability, consistency, and accessibility.",
    "An internal coding test uses \"{core}\" to verify whether candidates can convert requirements into a correct UI.",
    "In this practice track, \"{core}\" is scored on section ordering, spacing rhythm, and responsive behavior.",
    "A frontend screening module includes \"{core}\" to assess HTML semantics and robust CSS composition.",
    "As part of a UI implementation sprint, \"{core}\" must be completed with exact output and clean code conventions.",
    "A benchmark exercise uses \"{core}\" to check how well students recreate structured interfaces from specifications."
]

level_base = {
    'Beginner': [
        'Use only HTML and CSS; JavaScript is not allowed.',
        'Keep class names readable and beginner-friendly.',
        'Match spacing and font sizes close to the sample UI.',
    ],
    'Intermediate': [
        'Follow mobile-first CSS and include at least two breakpoints.',
        'Use reusable classes and avoid style duplication.',
        'Ensure visible keyboard focus on interactive controls.',
    ],
    'Hard': [
        'Support mobile, tablet, and desktop with stable layout behavior.',
        'Maintain accessibility contrast and consistent focus visibility.',
        'Use maintainable CSS architecture with predictable specificity.',
    ],
}

topic_rule = {
    'HTML': 'Use meaningful sectioning tags and valid document structure.',
    'Semantic HTML': 'Prioritize semantic landmarks instead of generic wrappers.',
    'Forms': 'Every input must have proper label association and logical grouping.',
    'Tables': 'Use caption, thead/tbody, and appropriate header scope attributes.',
    'CSS': 'Apply class-based styling; avoid inline style attributes.',
    'Box Model': 'Demonstrate intentional margin, padding, border, and sizing usage.',
    'Flexbox': 'Primary alignment/distribution should be solved with Flexbox properties.',
    'Grid': 'Main layout should be solved with CSS Grid tracks/areas.',
    'Media Queries': 'Breakpoint behavior must clearly alter layout/typography as required.',
    'Responsive Design': 'UI must adapt fluidly across screen widths without overlap/cutoff.',
}

extra_constraints = [
    'Section order must match the reference sequence exactly.',
    'Avoid hardcoded absolute positioning unless structurally necessary.',
    'No external UI framework usage is allowed for this task.',
    'Keep spacing increments consistent across repeated components.',
    'Do not omit any visible block shown in the expected output.',
    'Use semantic naming that reflects component purpose.',
    'Preserve readable line length and clear typography hierarchy.',
    'Ensure content blocks align cleanly on both small and large screens.',
    'Avoid unnecessary wrapper elements in markup.',
    'Submission should remain maintainable and easy to review.'
]


def choose(lst, key):
    h = int(hashlib.md5(key.encode()).hexdigest()[:8], 16)
    return lst[h % len(lst)]


def core_title(title):
    # remove trailing " - Topic Practice B001" etc
    return re.sub(r"\s-\s.+?Practice\s[BIH]\d{3}$", "", title).strip()

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        level = ws.cell(r, 2).value
        topic = ws.cell(r, 3).value
        old_title = ws.cell(r, 4).value
        core = core_title(old_title)
        ws.cell(r, 4).value = core

        key = f"{sheet}|{topic}|{core}|{r}"
        tech = 'HTML and CSS'
        ws.cell(r, 5).value = choose(desc_templates, key).format(core=core, topic=topic, tech=tech)
        ws.cell(r, 6).value = choose(scenario_templates, key + 'sc').format(core=core)

        c1 = choose(level_base[level], key + 'l1')
        c2 = topic_rule[topic]
        c3 = choose(extra_constraints, key + 'e1')
        ws.cell(r, 8).value = f"{c1} {c2} {c3}"

        # update linked html title/h1 to match updated question title
        link = ws.cell(r, 9).value
        if isinstance(link, str) and '/main/' in link:
            rel = link.split('/main/', 1)[1]
            html_path = repo / rel
            if html_path.exists() and html_path.suffix == '.html':
                s = html_path.read_text(encoding='utf-8')
                s = re.sub(r'<title>.*?</title>', f'<title>{core}</title>', s, count=1, flags=re.S)
                s = re.sub(r"<h1 class='title'>.*?</h1>", f"<h1 class='title'>{core}</h1>", s, count=1, flags=re.S)
                html_path.write_text(s, encoding='utf-8')

wb.save(wb_path)
print('Refined titles/descriptions/scenarios/constraints and synced output page titles.')
